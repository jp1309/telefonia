# -*- coding: utf-8 -*-
"""
Procesa: 1.1.1-Lineas-activas-por-servicio_y_Densidad_octubre_2025.xlsx
Hoja: "Líneas por servicio"
Datos: desde fila 79 (jul-2014) hacia abajo, hasta filas vacías (incluye futuras filas)

Salida:
- output/lineas_por_servicio_long.csv
- output/lineas_por_servicio_long.parquet (si tienes pyarrow)
- output/validaciones.csv
"""

from __future__ import annotations

import os
import re
import unicodedata
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# =========================
# CONFIG. AJUSTA SI QUIERES
# =========================
BASE_DIR = r"C:\Users\HP\OneDrive\JpE\Github\telefonia"
INPUT_XLSX = os.path.join(BASE_DIR, "1.1.1-Lineas-activas-por-servicio_y_Densidad_octubre_2025.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

SHEET_NAME = "Líneas por servicio"
HEADER_ROW = 11
START_ROW = 79

DATE_COL = "A"
TOTAL_MERCADO_COL = "Q"

COMPANY_BLOCKS = [
    ("CONECEL S.A.", ["B", "C", "D", "E"], "F"),
    ("OTECEL S.A.",  ["G", "H", "I", "J"], "K"),
    ("CNT EP",       ["L", "M", "N", "O"], "P"),
]

FALLBACK_START_DATE = date(2014, 7, 1)


def _normalize_text(x: str) -> str:
    x = x.strip().lower()
    x = "".join(c for c in unicodedata.normalize("NFKD", x) if not unicodedata.combining(c))
    x = re.sub(r"\s+", " ", x)
    return x


def _parse_month_year(raw) -> date | None:
    if raw is None:
        return None

    if hasattr(raw, "year") and hasattr(raw, "month"):
        try:
            return date(int(raw.year), int(raw.month), 1)
        except Exception:
            pass

    s = _normalize_text(str(raw))

    months = {
        "ene": 1, "enero": 1, "jan": 1, "january": 1,
        "feb": 2, "febrero": 2, "february": 2,
        "mar": 3, "marzo": 3, "march": 3,
        "abr": 4, "abril": 4, "apr": 4, "april": 4,
        "may": 5, "mayo": 5,
        "jun": 6, "junio": 6, "june": 6,
        "jul": 7, "julio": 7, "july": 7,
        "ago": 8, "agosto": 8, "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "septiembre": 9, "september": 9,
        "oct": 10, "octubre": 10, "october": 10,
        "nov": 11, "noviembre": 11, "november": 11,
        "dic": 12, "diciembre": 12, "dec": 12, "december": 12,
    }

    m = re.match(r"^([a-z]{3,9})(\d{4})$", s)
    if m and m.group(1) in months:
        return date(int(m.group(2)), months[m.group(1)], 1)

    m = re.search(r"([a-z]{3,12})\D+(\d{4})", s)
    if m and m.group(1) in months:
        return date(int(m.group(2)), months[m.group(1)], 1)

    m = re.search(r"([a-z]{3,12})\D+(\d{2})$", s)
    if m and m.group(1) in months:
        yy = int(m.group(2))
        year = 2000 + yy if yy <= 79 else 1900 + yy
        return date(year, months[m.group(1)], 1)

    m = re.search(r"(\d{1,2})\D+(\d{4})", s)
    if m:
        mm = int(m.group(1))
        yy = int(m.group(2))
        if 1 <= mm <= 12:
            return date(yy, mm, 1)

    m = re.search(r"(\d{4})\D+(\d{1,2})", s)
    if m:
        yy = int(m.group(1))
        mm = int(m.group(2))
        if 1 <= mm <= 12:
            return date(yy, mm, 1)

    return None


def _add_months(d: date, n: int) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)


def _cell(ws, col_letter: str, row: int):
    return ws.cell(row=row, column=column_index_from_string(col_letter)).value


def main():
    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(f"No encuentro el archivo: {INPUT_XLSX}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = load_workbook(INPUT_XLSX, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No encuentro la hoja '{SHEET_NAME}'. Hojas disponibles: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    services = []
    for c in COMPANY_BLOCKS[0][1]:  # B-E como canónico
        v = _cell(ws, c, HEADER_ROW)
        services.append(str(v).strip() if v is not None else c)

    raw_start = _cell(ws, DATE_COL, START_ROW)
    parsed_start = _parse_month_year(raw_start)
    start_date = parsed_start if parsed_start is not None else FALLBACK_START_DATE

    records = []
    date_quality_flags = []

    r = START_ROW
    month_idx = 0
    empty_streak = 0

    while empty_streak < 3 and r < START_ROW + 2000:
        raw_date = _cell(ws, DATE_COL, r)
        total_mercado = _cell(ws, TOTAL_MERCADO_COL, r)

        if raw_date is None and total_mercado is None:
            empty_streak += 1
            r += 1
            continue

        empty_streak = 0

        expected_date = _add_months(start_date, month_idx)
        parsed_raw = _parse_month_year(raw_date)

        date_quality_flags.append({
            "row_excel": r,
            "raw_date": raw_date,
            "expected_date": expected_date.isoformat(),
            "parsed_raw_date": parsed_raw.isoformat() if parsed_raw else None,
            "raw_date_unparseable": bool(parsed_raw is None),
            "raw_date_mismatch_expected": bool(parsed_raw is not None and parsed_raw != expected_date),
        })

        total_empresas = 0.0

        for company_name, service_cols, total_col in COMPANY_BLOCKS:
            company_service_sum = 0.0

            for svc_name, col_letter in zip(services, service_cols):
                val = _cell(ws, col_letter, r)
                if val is None:
                    continue
                try:
                    vnum = float(val)
                except Exception:
                    continue

                company_service_sum += vnum
                records.append({
                    "date": expected_date,
                    "company": company_name,
                    "category": svc_name,
                    "value": vnum,
                    "excel_row": r,
                })

            tot_val = _cell(ws, total_col, r)
            tot_num = None
            if tot_val is not None:
                try:
                    tot_num = float(tot_val)
                    total_empresas += tot_num
                except Exception:
                    tot_num = None

            records.append({"date": expected_date, "company": company_name, "category": "TOTAL_EMPRESA", "value": tot_num, "excel_row": r})
            records.append({"date": expected_date, "company": company_name, "category": "CHECK_SUM_SERVICIOS", "value": company_service_sum, "excel_row": r})

        mercado_num = None
        if total_mercado is not None:
            try:
                mercado_num = float(total_mercado)
            except Exception:
                mercado_num = None

        records.append({"date": expected_date, "company": "TOTAL_MERCADO", "category": "TOTAL_MERCADO", "value": mercado_num, "excel_row": r})
        records.append({"date": expected_date, "company": "TOTAL_MERCADO", "category": "CHECK_SUM_TOTALES_EMPRESA", "value": total_empresas, "excel_row": r})

        r += 1
        month_idx += 1

    df = pd.DataFrame(records)

    # Validaciones
    mercado = df[(df["company"] == "TOTAL_MERCADO") &
                 (df["category"].isin(["TOTAL_MERCADO", "CHECK_SUM_TOTALES_EMPRESA"]))] \
        .pivot_table(index=["date", "excel_row"], columns="category", values="value", aggfunc="first") \
        .reset_index()
    mercado["diff_mercado"] = mercado["TOTAL_MERCADO"] - mercado["CHECK_SUM_TOTALES_EMPRESA"]

    empresa_chk = df[df["category"].isin(["TOTAL_EMPRESA", "CHECK_SUM_SERVICIOS"])] \
        .pivot_table(index=["date", "excel_row", "company"], columns="category", values="value", aggfunc="first") \
        .reset_index()
    empresa_chk["diff_empresa"] = empresa_chk["TOTAL_EMPRESA"] - empresa_chk["CHECK_SUM_SERVICIOS"]

    fechas_chk = pd.DataFrame(date_quality_flags)

    # Salidas
    csv_path = os.path.join(OUTPUT_DIR, "lineas_por_servicio_long.csv")
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    parquet_path = os.path.join(OUTPUT_DIR, "lineas_por_servicio_long.parquet")
    parquet_ok = True
    try:
        df.to_parquet(parquet_path, index=False)
    except Exception:
        parquet_ok = False

    valid_path = os.path.join(OUTPUT_DIR, "validaciones.csv")
    mercado_out = mercado.assign(check="mercado")
    empresa_out = empresa_chk.assign(check="empresa")
    fechas_out = fechas_chk.assign(check="fechas")
    pd.concat([mercado_out, empresa_out, fechas_out], ignore_index=True).to_csv(valid_path, index=False, encoding="utf-8-sig")

    print("OK. Procesamiento terminado.")
    print(f"Archivo de entrada: {INPUT_XLSX}")
    print(f"Meses procesados: {month_idx}")
    print(f"CSV: {csv_path}")
    if parquet_ok:
        print(f"Parquet: {parquet_path}")
    else:
        print("Parquet no se generó. Falta pyarrow o hubo un error al escribirlo.")
    print(f"Validaciones: {valid_path}")

    bad_merc = mercado[mercado["diff_mercado"].abs() > 0.5]
    if not bad_merc.empty:
        print("\nALERTA. Diferencias en total mercado (Q vs suma empresas). Primeras filas:")
        print(bad_merc[["date", "excel_row", "diff_mercado"]].head(20).to_string(index=False))

    bad_emp = empresa_chk[empresa_chk["diff_empresa"].abs() > 0.5]
    if not bad_emp.empty:
        print("\nALERTA. Diferencias en totales por empresa (total vs suma servicios). Primeras filas:")
        print(bad_emp[["date", "excel_row", "company", "diff_empresa"]].head(20).to_string(index=False))

    fechas_warn = fechas_chk[(fechas_chk["raw_date_unparseable"]) | (fechas_chk["raw_date_mismatch_expected"])]
    if not fechas_warn.empty:
        print("\nNota. Fechas mal escritas o inconsistentes. Se corrigieron por secuencia mensual.")
        print(fechas_warn.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
