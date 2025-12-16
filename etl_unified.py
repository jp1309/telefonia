# -*- coding: utf-8 -*-
"""
ETL Unificado
Procesa ambos archivos de Excel para generar los CSVs necesarios para el dashboard.

1. Servicios: 1.1.1-Lineas-activas-por-servicio_y_Densidad_octubre_2025.xlsx
2. Modalidad: 1.1.2-Lineas-activas-por-modalidad_octubre_2025.xlsx

Salida:
- output/lineas_por_servicio_long.csv
- output/lineas_por_modalidad_fact.csv
- output/validaciones_unificadas.csv
"""

from __future__ import annotations

import os
import re
import unicodedata
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# =========================
# CONFIGURACIÓN GENERAL
# =========================
BASE_DIR = r"C:\Users\HP\OneDrive\JpE\Github\telefonia"
DOWNLOAD_DIR = os.path.join(BASE_DIR, "datos_descargados")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# Configuración de Estructura (Separado de los nombres de archivo)
# File 1 (Servicios)
FILE_1_SHEET = "Líneas por servicio"
FILE_1_HEADER_ROW = 11
FILE_1_START_ROW = 79
FILE_1_DATE_COL = "A"
FILE_1_TOTAL_COL = "Q"
FILE_1_BLOCKS = [
    ("CONECEL S.A.", ["B", "C", "D", "E"], "F"),
    ("OTECEL S.A.",  ["G", "H", "I", "J"], "K"),
    ("CNT EP",       ["L", "M", "N", "O"], "P"),
]
FILE_1_FALLBACK_DATE = date(2014, 7, 1)

# File 2 (Modalidad)
FILE_2_SHEET = "Lineas por modalidad"
FILE_2_HEADER_ROW = 12
FILE_2_START_ROW = 13
FILE_2_DATE_COL = "A"
FILE_2_TOTAL_COL = "Q"
FILE_2_BLOCKS = [
    ("CONECEL S.A.", ["B", "C", "D"], "E"),
    ("OTECEL S.A.",  ["F", "G", "H"], "I"),
    ("CNT EP",       ["J", "K", "L"], "M"),
]
FILE_2_FALLBACK_DATE = date(2008, 12, 1)

def detect_latest_files():
    """Busca en datos_descargados el par de archivos más reciente."""
    print(f"Buscando archivos recientes en: {DOWNLOAD_DIR}")
    
    if not os.path.exists(DOWNLOAD_DIR):
        print(f"ERROR: No existe el directorio {DOWNLOAD_DIR}")
        return None, None

    files = [f for f in os.listdir(DOWNLOAD_DIR) if f.lower().endswith('.xlsx')]
    
    # Regex para extraer fecha
    # Soporta: 1.1.1..._octubre_2025.xlsx o formatos similares
    mapa_fechas = {} # (anio, mes) -> {'servicio': path, 'modalidad': path}
    
    meses_es = {
        'ene': 1, 'enero': 1, 'feb': 2, 'febrero': 2, 'mar': 3, 'marzo': 3,
        'abr': 4, 'abril': 4, 'may': 5, 'mayo': 5, 'jun': 6, 'junio': 6,
        'jul': 7, 'julio': 7, 'ago': 8, 'agosto': 8, 'sep': 9, 'septiembre': 9,
        'oct': 10, 'octubre': 10, 'nov': 11, 'noviembre': 11, 'dic': 12, 'diciembre': 12
    }
    
    for f in files:
        path = os.path.join(DOWNLOAD_DIR, f)
        
        # Identificar tipo
        tipo = None
        if '1.1.1' in f or 'servicio' in f.lower():
            tipo = 'servicio'
        elif '1.1.2' in f or 'modalidad' in f.lower():
            tipo = 'modalidad'
        
        if not tipo: continue
        
        # Identificar fecha
        # Busca _Mes_Año ó Mes-Año
        match = re.search(r'[_\-\s]([a-zA-Z]+)[_\-\s]+(\d{4})', f)
        if not match:
             match = re.search(r'([a-zA-Z]+)-(\d{4})', f)
             
        if match:
            mes_str = match.group(1).lower()
            anio = int(match.group(2))
            mes_num = meses_es.get(mes_str, 0)
            
            if mes_num > 0:
                key = (anio, mes_num)
                if key not in mapa_fechas:
                    mapa_fechas[key] = {}
                mapa_fechas[key][tipo] = path

    # Buscar la fecha más alta que tenga ambos
    fechas_ordenadas = sorted(mapa_fechas.keys(), reverse=True)
    
    for fecha in fechas_ordenadas:
        if 'servicio' in mapa_fechas[fecha] and 'modalidad' in mapa_fechas[fecha]:
            print(f" - Localizado set completo para: {fecha[1]}/{fecha[0]}")
            return mapa_fechas[fecha]['servicio'], mapa_fechas[fecha]['modalidad']
            
    print("ERROR: No se encontró un par completo de archivos (1.1.1 y 1.1.2) para ninguna fecha.")
    return None, None


# =========================
# UTILIDADES COMPARTIDAS
# =========================

def _normalize_text(x: str) -> str:
    x = x.strip().lower()
    x = "".join(c for c in unicodedata.normalize("NFKD", x) if not unicodedata.combining(c))
    x = re.sub(r"\s+", " ", x)
    return x

def _parse_month_year(raw) -> date | None:
    if raw is None:
        return None

    if hasattr(raw, "year") and hasattr(raw, "month"):
        try:
            return date(int(raw.year), int(raw.month), 1)
        except Exception:
            pass

    s = _normalize_text(str(raw))

    months = {
        "ene": 1, "enero": 1, "jan": 1, "january": 1,
        "feb": 2, "febrero": 2, "february": 2,
        "mar": 3, "marzo": 3, "march": 3,
        "abr": 4, "abril": 4, "apr": 4, "april": 4,
        "may": 5, "mayo": 5,
        "jun": 6, "junio": 6, "june": 6,
        "jul": 7, "julio": 7, "july": 7,
        "ago": 8, "agosto": 8, "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "septiembre": 9, "september": 9,
        "oct": 10, "octubre": 10, "october": 10,
        "nov": 11, "noviembre": 11, "november": 11,
        "dic": 12, "diciembre": 12, "dec": 12, "december": 12,
    }

    # "Ene 2009", "October 2025"
    m = re.match(r"^([a-z]{3,9})(\d{4})$", s)
    if m and m.group(1) in months:
        return date(int(m.group(2)), months[m.group(1)], 1)
        
    m = re.search(r"([a-z]{3,12})\D+(\d{4})", s)
    if m and m.group(1) in months:
        return date(int(m.group(2)), months[m.group(1)], 1)

    m = re.search(r"([a-z]{3,12})\D+(\d{2})$", s)
    if m and m.group(1) in months:
        yy = int(m.group(2))
        year = 2000 + yy if yy <= 79 else 1900 + yy
        return date(year, months[m.group(1)], 1)

    # "07/2014"
    m = re.search(r"(\d{1,2})\D+(\d{4})", s)
    if m:
        mm = int(m.group(1))
        yy = int(m.group(2))
        if 1 <= mm <= 12:
            return date(yy, mm, 1)

    # "2014-07"
    m = re.search(r"(\d{4})\D+(\d{1,2})", s)
    if m:
        yy = int(m.group(1))
        mm = int(m.group(2))
        if 1 <= mm <= 12:
            return date(yy, mm, 1)

    return None

def _add_months(d: date, n: int) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)

def _cell(ws, col_letter: str, row: int):
    return ws.cell(row=row, column=column_index_from_string(col_letter)).value


# =========================
# LÓGICA DE PROCESAMIENTO
# =========================

def process_servicios(input_path):
    print(f"\n--- Procesando Servicios ---")
    print(f"Archivo: {os.path.basename(input_path)}")
    
    if not os.path.exists(input_path):
        print(f"ERROR: No se encontró {input_path}")
        return

    wb = load_workbook(input_path, data_only=True)
    if FILE_1_SHEET not in wb.sheetnames:
        print(f"ERROR: Hoja '{FILE_1_SHEET}' no encontrada.")
        return
    ws = wb[FILE_1_SHEET]

    # Detectar Nombres de Servicios (Headers)
    services = []
    # Usamos el bloque de CONECEL (B-E) para nombres canónicos
    for c in FILE_1_BLOCKS[0][1]:
        v = _cell(ws, c, FILE_1_HEADER_ROW)
        services.append(str(v).strip() if v is not None else c)
    
    # Fecha Inicio
    raw_start = _cell(ws, FILE_1_DATE_COL, FILE_1_START_ROW)
    parsed_start = _parse_month_year(raw_start)
    start_date = parsed_start if parsed_start is not None else FILE_1_FALLBACK_DATE

    records = []
    r = FILE_1_START_ROW
    month_idx = 0
    empty_streak = 0

    while empty_streak < 3 and r < FILE_1_START_ROW + 3000:
        raw_date = _cell(ws, FILE_1_DATE_COL, r)
        total_mercado = _cell(ws, FILE_1_TOTAL_COL, r)

        if raw_date is None and total_mercado is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0

        current_date = _add_months(start_date, month_idx)
        
        # Procesar bloques empresas
        sum_total_empresas = 0.0
        for company_name, service_cols, total_col in FILE_1_BLOCKS:
            company_sum = 0.0
            for svc_name, col_letter in zip(services, service_cols):
                val = _cell(ws, col_letter, r)
                vnum = 0.0
                if val is not None:
                    try: vnum = float(val)
                    except: pass
                
                company_sum += vnum
                records.append({
                    "date": current_date,
                    "company": company_name,
                    "category": svc_name,
                    "value": vnum,
                    "source": "servicios"
                })
            
            # Total empresa declarado
            tot_val = _cell(ws, total_col, r)
            tot_decl = 0.0
            if tot_val is not None:
                try: tot_decl = float(tot_val)
                except: pass
            sum_total_empresas += tot_decl

            records.append({"date": current_date, "company": company_name, "category": "TOTAL_EMPRESA", "value": tot_decl, "source": "servicios"})
            records.append({"date": current_date, "company": company_name, "category": "CHECK_SUM_SERVICIOS", "value": company_sum, "source": "servicios"})

        # Total Mercado
        merc_val = 0.0
        if total_mercado is not None:
            try: merc_val = float(total_mercado)
            except: pass
        
        records.append({"date": current_date, "company": "TOTAL_MERCADO", "category": "TOTAL_MERCADO", "value": merc_val, "source": "servicios"})
        records.append({"date": current_date, "company": "TOTAL_MERCADO", "category": "CHECK_SUM_TOTALES_EMPRESA", "value": sum_total_empresas, "source": "servicios"})

        r += 1
        month_idx += 1

    df = pd.DataFrame(records)
    out_path = os.path.join(OUTPUT_DIR, "lineas_por_servicio_long.csv")
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"Generado: {out_path} ({len(df)} registros)")

def process_modalidad(input_path):
    print(f"\n--- Procesando Modalidad ---")
    print(f"Archivo: {os.path.basename(input_path)}")
    
    if not os.path.exists(input_path):
        print(f"ERROR: No se encontró {input_path}")
        return

    wb = load_workbook(input_path, data_only=True)
    if FILE_2_SHEET not in wb.sheetnames:
        print(f"ERROR: Hoja '{FILE_2_SHEET}' no encontrada.")
        return
    ws = wb[FILE_2_SHEET]

    # Headers de Modalidades
    modalities = []
    for c in FILE_2_BLOCKS[0][1]:
        v = _cell(ws, c, FILE_2_HEADER_ROW)
        modalities.append(str(v).strip() if v is not None else c)

    # Fecha Inicio
    raw_start = _cell(ws, FILE_2_DATE_COL, FILE_2_START_ROW)
    parsed_start = _parse_month_year(raw_start)
    start_date = parsed_start if parsed_start is not None else FILE_2_FALLBACK_DATE

    records = []
    r = FILE_2_START_ROW
    month_idx = 0
    empty_streak = 0

    while empty_streak < 3 and r < FILE_2_START_ROW + 3000:
        raw_date = _cell(ws, FILE_2_DATE_COL, r)
        total_mercado = _cell(ws, FILE_2_TOTAL_COL, r)

        if raw_date is None and total_mercado is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0

        current_date = _add_months(start_date, month_idx)
        
        sum_total_empresas = 0.0
        for company_name, mod_cols, total_col in FILE_2_BLOCKS:
            company_sum = 0.0
            for mod_name, col_letter in zip(modalities, mod_cols):
                val = _cell(ws, col_letter, r)
                vnum = 0.0
                if val is not None:
                    try: vnum = float(val)
                    except: pass
                
                company_sum += vnum
                records.append({
                    "date": current_date,
                    "company": company_name,
                    "category": mod_name,
                    "value": vnum,
                    "source": "modalidad"
                })
            
            tot_val = _cell(ws, total_col, r)
            tot_decl = 0.0
            if tot_val is not None:
                try: tot_decl = float(tot_val)
                except: pass
            sum_total_empresas += tot_decl

            records.append({"date": current_date, "company": company_name, "category": "TOTAL_EMPRESA", "value": tot_decl, "source": "modalidad"})
            records.append({"date": current_date, "company": "CHECK_SUM_MODALIDADES", "value": company_sum, "source": "modalidad"})

        merc_val = 0.0
        if total_mercado is not None:
            try: merc_val = float(total_mercado)
            except: pass
            
        records.append({"date": current_date, "company": "TOTAL_MERCADO", "category": "TOTAL_MERCADO", "value": merc_val, "source": "modalidad"})
        records.append({"date": current_date, "company": "TOTAL_MERCADO", "category": "CHECK_SUM_TOTALES_EMPRESA", "value": sum_total_empresas, "source": "modalidad"})

        r += 1
        month_idx += 1

    df_long = pd.DataFrame(records)
    # Generar versión "Fact" limpia para dashboard (sin checks)
    df_fact = df_long[~df_long["category"].str.startswith("CHECK_")].copy()
    
    out_path_long = os.path.join(OUTPUT_DIR, "lineas_por_modalidad_long.csv")
    out_path_fact = os.path.join(OUTPUT_DIR, "lineas_por_modalidad_fact.csv")
    
    df_long.to_csv(out_path_long, index=False, encoding="utf-8-sig")
    df_fact.to_csv(out_path_fact, index=False, encoding="utf-8-sig")
    
    print(f"Generado: {out_path_long} ({len(df_long)} registros)")
    print(f"Generado: {out_path_fact} (Optimizado Dashboard)")

def main():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        
    print("Iniciando procesamiento ETL Auto-Detect...")
    
    f1, f2 = detect_latest_files()
    
    if f1 and f2:
        process_servicios(f1)
        process_modalidad(f2)
        print("\nPROCESO FINALIZADO EXITOSAMENTE.")
    else:
        print("\nABORTADO: No se pudieron determinar los archivos a procesar.")

if __name__ == "__main__":
    main()
